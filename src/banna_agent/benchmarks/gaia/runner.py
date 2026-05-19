"""GAIA benchmark runner.

One function — `run_gaia(tasks, *, llm, tools, policy, ...)` — drives each
`GAIATask` through the agent loop, scores the final answer, and returns a
`GAIARunResult` containing per-task outcomes plus aggregate summary.

It writes a per-task JSONL event log under `out_dir/logs/` and a single
`summary.json` + `results.jsonl` at the root. The layout makes week-2
ablation comparisons straightforward:

    runs/2026-05-06_react_haiku/
      summary.json            # aggregate scores + per-level breakdown
      results.jsonl           # one line per task
      logs/
        <task_id>.jsonl       # full event stream for replay

CLI is in `experiments/01_react_baseline/run.py` (created when that
experiment needs it). The function itself is library-only.
"""
from __future__ import annotations

import json
import time
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable, Sequence

from ...core.agent import run_policy
from ...core.events import EventLog
from ...core.state import AgentState
from ...core.types import Budget
from ...llm.base import LLMClient
from ...tools.base import ToolRegistry
from .loader import GAIATask
from .scorer import GAIAScore, score_gaia


# ---------------------------------------------------------------------------
# Per-task result
# ---------------------------------------------------------------------------


@dataclass
class GAIATaskResult:
    task_id: str
    level: int
    question: str
    gold_answer: str
    pred_answer: str
    is_correct: bool
    match_kind: str
    wall_s: float
    steps_used: int
    tokens_in: int
    tokens_out: int
    cost_usd: float
    budget_reason: str
    log_path: str
    error: str | None = None
    # Phase 1: structured ablation metrics. All optional/derived; older
    # consumers (summary aggregator) ignore them.
    policy_name: str = ""
    model_name: str = ""
    timeout: bool = False
    tool_calls_total: int = 0
    tool_calls_by_name: dict[str, int] = field(default_factory=dict)
    tool_errors_total: int = 0
    tool_errors_by_name: dict[str, int] = field(default_factory=dict)
    cache_hits: int = 0
    cache_misses: int = 0
    finished_reason: str = ""  # final | budget_steps | budget_wall | budget_cost | error

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


# ---------------------------------------------------------------------------
# Aggregate result
# ---------------------------------------------------------------------------


@dataclass
class GAIARunResult:
    n_total: int
    n_correct: int
    accuracy: float
    by_level: dict[int, dict[str, Any]] = field(default_factory=dict)
    by_match_kind: dict[str, dict[str, Any]] = field(default_factory=dict)
    total_tokens_in: int = 0
    total_tokens_out: int = 0
    total_cost_usd: float = 0.0
    total_wall_s: float = 0.0
    results: list[GAIATaskResult] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "n_total": self.n_total,
            "n_correct": self.n_correct,
            "accuracy": self.accuracy,
            "by_level": self.by_level,
            "by_match_kind": self.by_match_kind,
            "total_tokens_in": self.total_tokens_in,
            "total_tokens_out": self.total_tokens_out,
            "total_cost_usd": self.total_cost_usd,
            "total_wall_s": self.total_wall_s,
        }


# ---------------------------------------------------------------------------
# The driver
# ---------------------------------------------------------------------------


def run_gaia(
    tasks: Sequence[GAIATask],
    *,
    llm: LLMClient,
    tools: ToolRegistry,
    policy: Any,
    budget_factory: Any = None,
    out_dir: str | Path | None = None,
    verbose: bool = True,
    compactor: Any = None,
) -> GAIARunResult:
    """Drive every task through `policy` and score exact-match.

    Parameters
    ----------
    tasks           : iterable of `GAIATask`
    llm, tools      : pre-built LLM client and ToolRegistry
    policy          : an object with `.propose(state, *, llm, tools) -> Action`
    budget_factory  : callable(GAIATask) -> Budget; default caps at
                      level-appropriate step counts (L1=8, L2=14, L3=20)
                      and 120 s wall.
    out_dir         : optional directory to persist logs + summary
    verbose         : print per-task progress
    """
    if budget_factory is None:
        budget_factory = _default_budget

    out = Path(out_dir).expanduser().resolve() if out_dir else None
    if out is not None:
        (out / "logs").mkdir(parents=True, exist_ok=True)
        results_path = out / "results.jsonl"
        results_path.write_text("")  # truncate per-run

    results: list[GAIATaskResult] = []

    for i, task in enumerate(tasks, 1):
        t0 = time.monotonic()
        cache_baseline = _cache_stats_snapshot()
        log_path = None
        if out is not None:
            safe_id = task.task_id or f"task_{i}"
            safe_id = _safe_name(safe_id)
            log_path = out / "logs" / f"{safe_id}.jsonl"
        log = EventLog(log_path) if log_path else EventLog()

        state = AgentState(
            question=_format_question(task),
            budget=budget_factory(task),
            metadata={
                "task_id": task.task_id,
                "level": task.level,
                "has_attachment": task.has_attachment,
            },
        )

        err: str | None = None
        try:
            state = run_policy(state, policy, llm=llm, tools=tools, log=log,
                               compactor=compactor)
            pred = state.trace.final_answer or ""
        except Exception as exc:
            err = f"{type(exc).__name__}: {exc}"
            pred = ""

        # Post-run hook for policies that implement it. Errors here are
        # non-fatal — they don't change task scoring.
        if not err:
            post_hook = getattr(policy, "on_run_end", None)
            if post_hook is not None:
                try:
                    post_hook(state)
                except Exception:
                    pass

        score: GAIAScore = score_gaia(pred, task.answer)
        wall = time.monotonic() - t0
        reason = _last_budget_reason(log)

        tool_total, tool_by_name, tool_err_total, tool_err_by_name = _tool_call_stats(log)
        cache_h, cache_m = _cache_stats_delta(cache_baseline)

        finished_reason = (
            "error" if err else
            "final" if state.trace.final_answer else
            reason if reason and reason != "ok" else
            "no_final_answer"
        )

        tr = GAIATaskResult(
            task_id=task.task_id,
            level=task.level,
            question=task.question,
            gold_answer=task.answer,
            pred_answer=pred,
            is_correct=score.is_correct,
            match_kind=score.match_kind,
            wall_s=wall,
            steps_used=state.budget.steps_used,
            tokens_in=state.budget.tokens_in,
            tokens_out=state.budget.tokens_out,
            cost_usd=state.budget.cost_usd,
            budget_reason=reason,
            log_path=str(log_path) if log_path else "",
            error=err,
            policy_name=type(policy).__name__,
            model_name=getattr(llm, "model", "") or getattr(llm, "model_name", ""),
            timeout=reason.startswith("budget_wall") or reason == "timeout",
            tool_calls_total=tool_total,
            tool_calls_by_name=tool_by_name,
            tool_errors_total=tool_err_total,
            tool_errors_by_name=tool_err_by_name,
            cache_hits=cache_h,
            cache_misses=cache_m,
            finished_reason=finished_reason,
        )
        results.append(tr)

        if out is not None:
            with (out / "results.jsonl").open("a", encoding="utf-8") as f:
                f.write(json.dumps(tr.to_dict(), default=str) + "\n")

        if verbose:
            mark = "✓" if tr.is_correct else "✗"
            print(
                f"[{i:>3}/{len(tasks)}] L{task.level} {mark} "
                f"{task.task_id or '-':<10} "
                f"pred={_truncate(pred, 40)!r:<42} "
                f"gold={_truncate(task.answer, 40)!r} "
                f"({tr.steps_used}st {wall:.1f}s)"
            )

    agg = _aggregate(results)

    if out is not None:
        (out / "summary.json").write_text(json.dumps(agg.to_dict(), indent=2, default=str))

    return agg


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _default_budget(task: GAIATask) -> Budget:
    # Bumped from {1:8, 2:14, 3:20}. The 05-16 gpt-5-nano full-validation
    # run had 19/27 budget_steps failures on L1 with steps=8 — verifier
    # nudges and commit-required nudges each ate a step, leaving 5–6
    # productive steps for the actual task. Repair steps are now
    # exempt from this count (see Budget.max_repair_steps + the
    # `repair=True` action.meta tag in policies/react.py), but a fatter
    # main budget still gives the model room to backtrack on hard L2/L3
    # tasks that legitimately need 10+ tool calls.
    steps = {1: 12, 2: 18, 3: 24}.get(task.level, 14)
    return Budget(max_steps=steps, max_wall_s=120.0)


_RICH_TOOL_HINTS: dict[str, str] = {
    ".pdf":  "pdf_open / pdf_read_page / pdf_find / pdf_read_tables",
    ".xlsx": "xlsx_list_sheets / xlsx_describe / xlsx_read_range / xlsx_find",
    ".xls":  "xlsx_list_sheets / xlsx_describe / xlsx_read_range / xlsx_find",
    ".docx": "read_file (auto-extracts text from the .docx zip)",
}
_TEXT_EXTS = {".csv", ".txt", ".json", ".md", ".tsv", ".log", ".py", ".yaml", ".yml"}


def _file_hint(path: str) -> str:
    """Extension-routed hint pointing at the right tool family.

    The pre-C5 hint only mentioned `read_file` / `run_python`, which
    actively steered models away from the structured pdf_* / xlsx_*
    tools. Across the 05-16 gpt-5-nano full-validation run that meant
    0 pdf_open calls and 1 xlsx_* call across 38 attachment tasks.
    """
    ext = Path(path).suffix.lower()
    if ext in _RICH_TOOL_HINTS:
        return f"For this file ({ext}), prefer: {_RICH_TOOL_HINTS[ext]}."
    if ext in _TEXT_EXTS:
        return "Use `read_file`; for large/structured cases use `run_python`."
    return (
        "Use `read_file` first to sniff the contents; use `run_python` "
        "for binary parsing."
    )


def _file_summary(path: str) -> str:
    """Cheap one-line metadata: page count / sheet names / csv header.

    Failures are silent — if introspection breaks we drop the line.
    Capped at 80 chars so the prompt stays compact.
    """
    try:
        p = Path(path)
        if not p.exists() or p.is_dir():
            return ""
        ext = p.suffix.lower()
        if ext == ".pdf":
            try:
                from pypdf import PdfReader  # type: ignore
                n = len(PdfReader(str(p)).pages)
                return f"File summary: PDF with {n} page(s)."
            except Exception:
                return ""
        if ext in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook  # type: ignore
                wb = load_workbook(str(p), read_only=True, data_only=True)
                names = wb.sheetnames[:5]
                more = "" if len(wb.sheetnames) <= 5 else f" (+{len(wb.sheetnames)-5} more)"
                return f"File summary: sheets {names}{more}."
            except Exception:
                return ""
        if ext in (".csv", ".tsv"):
            try:
                first = p.open("r", errors="replace").readline().rstrip("\n")
                if len(first) > 70:
                    first = first[:67] + "…"
                return f"File summary: header line — {first}"
            except Exception:
                return ""
        return ""
    except Exception:
        return ""


def _format_question(task: GAIATask) -> str:
    prompt = task.question
    if task.has_attachment and task.file_path:
        lines = [
            prompt,
            "",
            f"An attached file is available at: {task.file_path}",
            _file_hint(task.file_path),
        ]
        summary = _file_summary(task.file_path)
        if summary:
            lines.append(summary)
        prompt = "\n".join(lines)
    prompt += (
        "\n\nAnswer with just the short answer — a number, a single name, "
        "or a comma-separated list as appropriate. No preamble, no "
        "explanation, no units unless the question asks for them."
    )
    return prompt


def _safe_name(s: str) -> str:
    return "".join(c if c.isalnum() or c in "-_." else "_" for c in s)[:80]


def _truncate(s: str, n: int) -> str:
    s = s.replace("\n", " ")
    return s if len(s) <= n else s[: n - 1] + "…"


def _cache_stats_snapshot() -> tuple[int, int]:
    """Return (hits, misses) on the current HTTP cache singleton.

    Wrapped in a try/except so a missing/disabled cache never breaks a
    GAIA run. Returns (0, 0) when the cache module isn't importable or
    the singleton is in `off` mode (we still read the counters; they're
    just zero).
    """
    try:
        from ...tools._http_cache import get_cache
        s = get_cache().stats()
        return s.get("hits", 0), s.get("misses", 0)
    except Exception:
        return 0, 0


def _cache_stats_delta(baseline: tuple[int, int]) -> tuple[int, int]:
    h0, m0 = baseline
    h1, m1 = _cache_stats_snapshot()
    return max(0, h1 - h0), max(0, m1 - m0)


def _tool_call_stats(log: EventLog) -> tuple[int, dict[str, int], int, dict[str, int]]:
    """Count tool_call/tool_result events on the log.

    Returns (total_calls, calls_by_name, total_errors, errors_by_name).
    The event schema (see core/agent.py): `tool_call` events carry
    ``tool_name``; `tool_result` events carry ``ok``/``error`` and are
    matched back to their `tool_call` by step index.
    """
    total = 0
    by_name: dict[str, int] = {}
    err_total = 0
    err_by_name: dict[str, int] = {}
    pending: dict[int, str] = {}  # step -> tool name

    for ev in log.events:
        if ev.kind == "tool_call":
            name = str(ev.payload.get("tool_name") or "?")
            total += 1
            by_name[name] = by_name.get(name, 0) + 1
            pending[ev.step] = name
        elif ev.kind == "tool_result":
            name = pending.pop(ev.step, None) or "?"
            ok = ev.payload.get("ok")
            had_error = (ok is False) or bool(ev.payload.get("error"))
            if had_error:
                err_total += 1
                err_by_name[name] = err_by_name.get(name, 0) + 1
    return total, by_name, err_total, err_by_name


def _last_budget_reason(log: EventLog) -> str:
    for ev in reversed(log.events):
        if ev.kind == "run_end":
            return ev.payload.get("budget_reason", "ok")
        if ev.kind == "budget":
            return ev.payload.get("reason", "ok")
    return "ok"


def _aggregate(results: list[GAIATaskResult]) -> GAIARunResult:
    total = len(results)
    correct = sum(1 for r in results if r.is_correct)

    by_level: dict[int, dict[str, Any]] = {}
    for r in results:
        slot = by_level.setdefault(r.level, {"n_total": 0, "n_correct": 0})
        slot["n_total"] += 1
        if r.is_correct:
            slot["n_correct"] += 1
    for lv, s in by_level.items():
        s["accuracy"] = s["n_correct"] / s["n_total"] if s["n_total"] else 0.0

    by_kind: dict[str, dict[str, Any]] = {}
    for r in results:
        slot = by_kind.setdefault(r.match_kind, {"n_total": 0, "n_correct": 0})
        slot["n_total"] += 1
        if r.is_correct:
            slot["n_correct"] += 1
    for k, s in by_kind.items():
        s["accuracy"] = s["n_correct"] / s["n_total"] if s["n_total"] else 0.0

    return GAIARunResult(
        n_total=total,
        n_correct=correct,
        accuracy=(correct / total) if total else 0.0,
        by_level=by_level,
        by_match_kind=by_kind,
        total_tokens_in=sum(r.tokens_in for r in results),
        total_tokens_out=sum(r.tokens_out for r in results),
        total_cost_usd=sum(r.cost_usd for r in results),
        total_wall_s=sum(r.wall_s for r in results),
        results=results,
    )
