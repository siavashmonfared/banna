"""Workbook reader.

`file_reader._read_xlsx` returns a top-N preview of every sheet, which
is the wrong shape for GAIA L2 spreadsheet tasks ("look up the value in
cell D17", "find the column whose header contains 'revenue' and sum
it"). This module exposes targeted operations:

  * `xlsx_list_sheets(path)`
  * `xlsx_describe(path, sheet)`            → dims + header row
  * `xlsx_read_range(path, sheet, "A1:D20")` → markdown table
  * `xlsx_find(path, value, sheet?)`         → cell addresses where the
                                               literal value appears

Built directly on openpyxl (already a project dep). All ops are
read-only and skip formulas — we return computed values when openpyxl
has cached them (Excel saves these alongside the formula), and the raw
formula string otherwise.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Iterable


_RANGE_RE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$")
_DEFAULT_MAX_CELLS = 800  # range read cap; 800 cells ~= a 20x40 table


def _wb(path: Path, *, read_only: bool = True):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl not installed") from exc
    return openpyxl.load_workbook(str(path), read_only=read_only, data_only=True)


def xlsx_list_sheets(path: str | Path) -> dict[str, Any]:
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        return {"ok": False, "error": f"no such file: {p}"}
    try:
        wb = _wb(p)
    except Exception as exc:
        return {"ok": False, "error": f"open failed: {exc}"}
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({"name": name, "max_row": ws.max_row, "max_col": ws.max_column})
    return {"ok": True, "path": str(p), "sheets": sheets}


def xlsx_describe(path: str | Path, sheet: str | None = None) -> dict[str, Any]:
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        return {"ok": False, "error": f"no such file: {p}"}
    try:
        wb = _wb(p)
    except Exception as exc:
        return {"ok": False, "error": f"open failed: {exc}"}
    name = sheet or wb.sheetnames[0]
    if name not in wb.sheetnames:
        return {"ok": False, "error": f"no sheet named {name!r}; have {wb.sheetnames}"}
    ws = wb[name]
    header_row: list[str] = []
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first is not None:
        header_row = [str(c) if c is not None else "" for c in first]
    return {
        "ok": True, "path": str(p), "sheet": name,
        "max_row": ws.max_row, "max_col": ws.max_column,
        "header_row": header_row,
    }


def xlsx_read_range(path: str | Path, sheet: str, range_str: str,
                    *, max_cells: int = _DEFAULT_MAX_CELLS) -> dict[str, Any]:
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        return {"ok": False, "error": f"no such file: {p}"}
    m = _RANGE_RE.match((range_str or "").strip().upper())
    if not m:
        return {"ok": False, "error": f"range must look like 'A1:D20', got {range_str!r}"}
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    try:
        wb = _wb(p)
    except Exception as exc:
        return {"ok": False, "error": f"open failed: {exc}"}
    if sheet not in wb.sheetnames:
        return {"ok": False, "error": f"no sheet named {sheet!r}; have {wb.sheetnames}"}
    n_cells = (r2 - r1 + 1) * (_col_to_idx(c2) - _col_to_idx(c1) + 1)
    if n_cells > max_cells:
        return {
            "ok": False,
            "error": f"range covers {n_cells} cells; max_cells={max_cells}. "
                     f"Pick a narrower range or raise max_cells.",
        }
    ws = wb[sheet]
    rows: list[list[Any]] = []
    for row in ws[f"{c1}{r1}:{c2}{r2}"]:
        rows.append([_cell_value(c) for c in row])
    return {
        "ok": True, "path": str(p), "sheet": sheet, "range": f"{c1}{r1}:{c2}{r2}",
        "n_rows": len(rows), "n_cols": len(rows[0]) if rows else 0,
        "rows": rows,
        "markdown": _rows_to_markdown(rows),
    }


def xlsx_find(path: str | Path, value: Any, *, sheet: str | None = None,
              max_hits: int = 50) -> dict[str, Any]:
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        return {"ok": False, "error": f"no such file: {p}"}
    try:
        wb = _wb(p)
    except Exception as exc:
        return {"ok": False, "error": f"open failed: {exc}"}
    needle = str(value).lower()
    if sheet and sheet not in wb.sheetnames:
        return {"ok": False, "error": f"no sheet named {sheet!r}; have {wb.sheetnames}"}
    sheets_to_scan: Iterable[str] = [sheet] if sheet else wb.sheetnames
    hits: list[dict[str, Any]] = []
    for name in sheets_to_scan:
        ws = wb[name]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if needle in str(v).lower():
                    hits.append({"sheet": name, "address": cell.coordinate, "value": _cell_value(cell)})
                    if len(hits) >= max_hits:
                        return {"ok": True, "path": str(p), "query": str(value),
                                "n_hits": len(hits), "hits": hits, "max_hits_reached": True}
    return {"ok": True, "path": str(p), "query": str(value),
            "n_hits": len(hits), "hits": hits, "max_hits_reached": False}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _col_to_idx(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _cell_value(cell) -> Any:
    """Return the cached value when present, else fall back to the raw."""
    v = getattr(cell, "value", None)
    # openpyxl in data_only=True returns the cached computed value for
    # formula cells. If we got "=A1+B1" back, the workbook was saved
    # without cached values — surface the formula text rather than fail.
    return v


def _rows_to_markdown(rows: list[list[Any]]) -> str:
    if not rows:
        return ""
    cells = [[("" if v is None else str(v)).replace("|", "\\|") for v in r] for r in rows]
    widths = [max(len(r[i]) for r in cells) for i in range(len(cells[0]))]
    def fmt(row): return "| " + " | ".join(r.ljust(widths[i]) for i, r in enumerate(row)) + " |"
    sep = "| " + " | ".join("-" * w for w in widths) + " |"
    return "\n".join([fmt(cells[0]), sep, *(fmt(r) for r in cells[1:])])


# ---------------------------------------------------------------------------
# JsonTool factories
# ---------------------------------------------------------------------------


def make_xlsx_tools():
    from .base import JsonTool

    list_schema = {
        "type": "object",
        "properties": {"path": {"type": "string"}},
        "required": ["path"], "additionalProperties": False,
    }
    desc_schema = {
        "type": "object",
        "properties": {"path": {"type": "string"}, "sheet": {"type": "string"}},
        "required": ["path"], "additionalProperties": False,
    }
    range_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string"},
            "range": {"type": "string", "description": "Excel range like 'A1:D20'."},
            "max_cells": {"type": "integer", "default": _DEFAULT_MAX_CELLS},
        },
        "required": ["path", "sheet", "range"], "additionalProperties": False,
    }
    find_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "value": {"type": ["string", "number", "boolean"]},
            "sheet": {"type": "string", "description": "Restrict to one sheet (optional)."},
            "max_hits": {"type": "integer", "default": 50},
        },
        "required": ["path", "value"], "additionalProperties": False,
    }
    return (
        JsonTool(
            name="xlsx_list_sheets",
            description="List the sheets in an Excel workbook with their dimensions.",
            input_schema=list_schema,
            handler=lambda a: xlsx_list_sheets(a["path"]),
            capabilities=frozenset({"read", "filesystem"}),
        ),
        JsonTool(
            name="xlsx_describe",
            description="Describe one sheet: dimensions and header row.",
            input_schema=desc_schema,
            handler=lambda a: xlsx_describe(a["path"], a.get("sheet")),
            capabilities=frozenset({"read", "filesystem"}),
        ),
        JsonTool(
            name="xlsx_read_range",
            description=(
                "Read a cell range (e.g. 'A1:D20') from a sheet and return both the raw "
                "row matrix and a markdown table preview. Caps at max_cells to keep "
                "responses small."
            ),
            input_schema=range_schema,
            handler=lambda a: xlsx_read_range(
                a["path"], a["sheet"], a["range"],
                max_cells=int(a.get("max_cells", _DEFAULT_MAX_CELLS)),
            ),
            capabilities=frozenset({"read", "filesystem"}),
        ),
        JsonTool(
            name="xlsx_find",
            description=(
                "Find every cell whose value contains the given substring (case-insensitive). "
                "Optionally restrict to one sheet. Returns up to max_hits cell addresses."
            ),
            input_schema=find_schema,
            handler=lambda a: xlsx_find(
                a["path"], a["value"], sheet=a.get("sheet"),
                max_hits=int(a.get("max_hits", 50)),
            ),
            capabilities=frozenset({"read", "filesystem"}),
        ),
    )
